import os
import re
import pandas as pd
from datetime import datetime, timedelta
from data_processor import DataProcessor
from pivot_generator import generate_execute_period_pivot
from openpyxl import load_workbook

TOTAL_OHT_COUNT = 69

def process_transfer_time_excel(file_path: str) -> pd.DataFrame:
    processor = DataProcessor(file_path)
    processor.load_sheets()
    processor.enrich_data()
    df = processor.get_transformed_data()
    df['EXECUTE PERIOD'] = pd.to_numeric(df['EXECUTE PERIOD'], errors='coerce')
    df = df.dropna(subset=['EXECUTE PERIOD'])
    return df

def compute_hourly_avg(df: pd.DataFrame) -> pd.DataFrame:
    hourly_avg = df.groupby('Hour')['EXECUTE PERIOD'].mean().reset_index()
    hourly_avg.columns = ['Hour', 'Avg_EXECUTE_PERIOD']
    return hourly_avg

def compute_oht_utilization(df: pd.DataFrame, total_oht: int = TOTAL_OHT_COUNT) -> pd.DataFrame:
    used_oht_count = df['OHT ID'].nunique()
    utilization_percent = round((used_oht_count / total_oht) * 100, 2)
    return pd.DataFrame([{
        'Used OHTs': used_oht_count,
        'Total OHTs': total_oht,
        'Utilization (%)': utilization_percent
    }])

def update_weekly_report(output_file: str, date_tag: str):
    wb = load_workbook(output_file)

    processed_sheet = f"{date_tag}_Processed"
    hourly_sheet = f"{date_tag}_HourlyADT"
    utilization_sheet = f"{date_tag}_Utilization"
    weekly_sheet = "Weekly_Report"

    required_sheets = [processed_sheet, hourly_sheet, utilization_sheet]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        print(f"❌ Skipping Weekly_Report update — missing sheet(s): {missing}")
        return

    df_processed = pd.read_excel(output_file, sheet_name=processed_sheet)
    df_hourly = pd.read_excel(output_file, sheet_name=hourly_sheet)
    df_utilization = pd.read_excel(output_file, sheet_name=utilization_sheet)

    avg_exec_period = df_hourly['Avg_EXECUTE_PERIOD'].mean()
    transfer_count = len(df_processed)
    utilization_percent = df_utilization.iloc[0]['Utilization (%)']

    today_obj = datetime.strptime(date_tag, "%Y%m%d").date()

    if weekly_sheet in wb.sheetnames:
        df_weekly = pd.read_excel(output_file, sheet_name=weekly_sheet)
    else:
        df_weekly = pd.DataFrame(columns=[
            'Date', 'Avg EXECUTE PERIOD', 'OHT Utilization (%)', 'Failure Rate (%)', 'Transfer Count'
        ])

    if 'Date' not in df_weekly.columns:
        df_weekly['Date'] = ""

    exists = any(
        pd.to_datetime(row, errors='coerce').date() == today_obj
        for row in df_weekly['Date'] if row != 'Weekly Avg'
    )

    if exists:
        print(f"✅ {today_obj} already exists in Weekly_Report — skipped.")
        return

    print(f"📌 Inserting {today_obj} into Weekly_Report")

    df_data = df_weekly[df_weekly['Date'] != 'Weekly Avg'].copy()

    # ➕ Add new row with formatted %
    new_row = pd.DataFrame([{
        'Date': today_obj,
        'Avg EXECUTE PERIOD': avg_exec_period,
        'OHT Utilization (%)': f"{utilization_percent:.2f}%",
        'Failure Rate (%)': pd.NA,
        'Transfer Count': transfer_count
    }])

    # 📊 Prepare for concat & sorting
    df_data['SortDate'] = pd.to_datetime(df_data['Date'], errors='coerce')
    df_data = df_data.astype(new_row.dtypes.to_dict(), errors='ignore')
    df_new = pd.concat([df_data, new_row], ignore_index=True)
    df_new = df_new.sort_values(by='SortDate', na_position='last').drop(columns=['SortDate'])

    # 📈 Recompute Weekly Avg
    try:
        util_vals = df_new['OHT Utilization (%)'].dropna().apply(lambda x: float(str(x).replace('%', '').strip()))
        util_avg = util_vals.mean()
    except Exception:
        util_avg = pd.NA

    try:
        fail_vals = df_new['Failure Rate (%)'].dropna().apply(lambda x: float(str(x).replace('%', '').strip()))
        fail_avg = f"{fail_vals.mean():.5f}%" if not fail_vals.empty else pd.NA
    except Exception:
        fail_avg = pd.NA

    df_avg = {
        'Date': 'Weekly Avg',
        'Avg EXECUTE PERIOD': df_new['Avg EXECUTE PERIOD'].apply(pd.to_numeric, errors='coerce').mean(),
        'OHT Utilization (%)': f"{util_avg:.2f}%" if pd.notna(util_avg) else pd.NA,
        'Failure Rate (%)': fail_avg,
        'Transfer Count': df_new['Transfer Count'].sum()
    }

    df_final = pd.concat([df_new, pd.DataFrame([df_avg])], ignore_index=True)

    # 💾 Write Weekly_Report
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_final.to_excel(writer, sheet_name=weekly_sheet, index=False)

    # 🔚 Move Weekly_Report to end
    wb = load_workbook(output_file)
    if weekly_sheet in wb.sheetnames:
        ws = wb[weekly_sheet]
        wb._sheets.remove(ws)
        wb._sheets.append(ws)
        wb.save(output_file)
        print(f"✅ Weekly_Report updated and moved to the end.")
    wb.close()


if __name__ == "__main__":
    data_dir = "data"
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)

    transfer_file = next((f for f in os.listdir(data_dir) if f.startswith("TransferTime")), None)
    if not transfer_file:
        raise FileNotFoundError("❌ No file starting with 'TransferTime' found.")

    input_path = os.path.join(data_dir, transfer_file)
    output_file = os.path.join(output_dir, "OHT_Daily_Report.xlsx")
    file_exists = os.path.exists(output_file)

    # 📅 Extract date from filename
    match = re.search(r'20\d{6}', transfer_file)
    if match:
        file_date = datetime.strptime(match.group(), '%Y%m%d')
        date_tag = (file_date - timedelta(days=1)).strftime('%Y%m%d')
        print(f"📆 Detected date in filename: {match.group()} → Using date_tag: {date_tag}")
    else:
        date_tag = (datetime.today() - timedelta(days=1)).strftime('%Y%m%d')
        print(f"⚠️ No valid date found in filename — defaulting to: {date_tag}")

    df_result = process_transfer_time_excel(input_path)
    pivot_df = generate_execute_period_pivot(df_result)
    hourly_df = compute_hourly_avg(df_result)
    utilization_df = compute_oht_utilization(df_result)

    # ✅ Handle mode and if_sheet_exists based on file existence
    mode = 'a' if file_exists else 'w'
    writer_kwargs = {
        "engine": "openpyxl",
        "mode": mode
    }
    if mode == 'a':
        writer_kwargs["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(output_file, **writer_kwargs) as writer:
        df_result.to_excel(writer, sheet_name=f'{date_tag}_Processed', index=False)
        pivot_df.to_excel(writer, sheet_name=f'{date_tag}_PivotSource', index=False)
        hourly_df.to_excel(writer, sheet_name=f'{date_tag}_HourlyADT', index=False)
        utilization_df.to_excel(writer, sheet_name=f'{date_tag}_Utilization', index=False)

    update_weekly_report(output_file, date_tag)
